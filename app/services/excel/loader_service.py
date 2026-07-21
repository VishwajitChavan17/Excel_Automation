"""
app.services.excel.loader_service
===================================
Loads xlsx / xls / xlsm / csv / tsv / ods files and produces a
WorkbookHandle plus per-column profiling statistics, and every sheet as a
DataFrame (needed for multi-sheet preview tabs and cross-sheet tool
selection in Compare / Lookup / Duplicate Finder).

Engine selection strategy:
- openpyxl (via pandas) is used to enumerate sheets and read small/medium
  files, because it correctly handles multi-sheet xlsx, hidden sheets, and
  merged cells.
- For files whose row count exceeds LARGE_FILE_ROW_THRESHOLD, the CSV/TSV
  path switches to Polars, which is dramatically faster and more memory
  efficient for large flat files (500k+ rows).

This module has no Qt dependency -- it is pure logic, callable from a
QThread worker, a CLI, or a unit test without pulling in the GUI stack.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
import polars as pl
from loguru import logger

from app.core import constants
from app.services.excel.models import ColumnProfile, SheetInfo, WorkbookHandle


class UnsupportedFileTypeError(ValueError):
    pass


def _sheet_names_and_visibility(file_path: Path) -> dict[str, bool]:
    """Return {sheet_name: is_hidden} using openpyxl in read-only/metadata
    mode, without loading full cell data."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    try:
        return {ws.title: ws.sheet_state != "visible" for ws in wb.worksheets}
    finally:
        wb.close()


def profile_dataframe(df: pd.DataFrame) -> list[ColumnProfile]:
    """Compute per-column statistics used by the Right Panel / Properties view."""
    profiles: list[ColumnProfile] = []
    row_count = len(df) if len(df) else 1  # avoid div-by-zero on empty sheets

    for col in df.columns:
        series = df[col]
        null_count = int(series.isna().sum())
        unique_count = int(series.nunique(dropna=True))
        duplicate_count = max(0, (row_count - null_count) - unique_count)

        # Only need a handful of example values, so bound the (relatively
        # expensive, since it's a full string cast) scan to a small sample
        # rather than casting every non-null value in the column. Confirmed
        # via profiling this was the dominant per-column cost for
        # high-cardinality numeric columns at scale (~0.58s for 500k mostly-
        # unique floats before this fix).
        non_null = series.dropna()
        sample = non_null.head(200) if len(non_null) > 200 else non_null
        example_values = sample.astype(str).unique()[:5].tolist() if row_count else []

        lengths = series.dropna().astype(str).map(len)
        max_len = int(lengths.max()) if not lengths.empty else None
        min_len = int(lengths.min()) if not lengths.empty else None

        is_numeric = pd.api.types.is_numeric_dtype(series)
        min_value = max_value = None
        if is_numeric and not series.dropna().empty:
            raw_min, raw_max = series.min(), series.max()
            min_value = float(raw_min) if pd.notna(raw_min) else None
            max_value = float(raw_max) if pd.notna(raw_max) else None

        profiles.append(
            ColumnProfile(
                name=str(col),
                dtype=str(series.dtype),
                null_count=null_count,
                null_pct=round(100 * null_count / row_count, 2),
                unique_count=unique_count,
                unique_pct=round(100 * unique_count / row_count, 2),
                duplicate_pct=round(100 * duplicate_count / row_count, 2),
                example_values=example_values,
                max_length=max_len,
                min_length=min_len,
                min_value=min_value,
                max_value=max_value,
                is_numeric=is_numeric,
            )
        )
    return profiles


def _row_level_stats(df: pd.DataFrame) -> tuple[int, int]:
    """Return (full_row_duplicate_count, total_blank_cell_count)."""
    duplicate_row_count = int(df.duplicated(keep="first").sum()) if len(df) else 0
    blank_cell_count = int(df.isna().sum().sum())
    return duplicate_row_count, blank_cell_count


def load_workbook(
    file_path: str | Path,
    *,
    sheet_name: str | None = None,
    large_file_threshold: int = constants.LARGE_FILE_ROW_THRESHOLD,
) -> tuple[WorkbookHandle, pd.DataFrame]:
    """Load a single file and return (WorkbookHandle, DataFrame-for-active-sheet).
    Convenience wrapper around load_workbook_all_sheets() for callers (and
    tests) that only care about one sheet."""
    handle, sheets = load_workbook_all_sheets(
        file_path, sheet_name=sheet_name, large_file_threshold=large_file_threshold
    )
    return handle, sheets[handle.active_sheet]


def load_workbook_all_sheets(
    file_path: str | Path,
    *,
    sheet_name: str | None = None,
    large_file_threshold: int = constants.LARGE_FILE_ROW_THRESHOLD,
) -> tuple[WorkbookHandle, dict[str, pd.DataFrame]]:
    """
    Load every sheet in a workbook (a single "sheet" for CSV/TSV) and return
    (WorkbookHandle, {sheet_name: DataFrame}). The handle's column_profiles /
    duplicate_row_count / blank_cell_count describe `active_sheet` only;
    switching the active sheet in the UI re-profiles from the cached frame
    without touching disk again.

    Raises:
        FileNotFoundError: path does not exist.
        UnsupportedFileTypeError: extension not in SUPPORTED_EXCEL_EXTENSIONS.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(path)

    ext = path.suffix.lower()
    if ext not in constants.SUPPORTED_EXCEL_EXTENSIONS:
        raise UnsupportedFileTypeError(f"Unsupported file type: {ext}")

    logger.info("Loading workbook: {}", path)
    stat = path.stat()

    if ext in (".csv", ".tsv"):
        handle, sheets = _load_delimited(path, ext, large_file_threshold)
    elif ext == ".ods":
        handle, sheets = _load_with_pandas(path, sheet_name, engine="odf")
    else:  # .xlsx, .xlsm, .xls
        handle, sheets = _load_with_pandas(path, sheet_name, engine=None)

    handle.file_size_bytes = stat.st_size
    handle.last_modified = datetime.fromtimestamp(stat.st_mtime)

    active_df = sheets[handle.active_sheet]
    handle.column_profiles = profile_dataframe(active_df)
    handle.duplicate_row_count, handle.blank_cell_count = _row_level_stats(active_df)

    return handle, sheets


def reprofile_for_sheet(handle: WorkbookHandle, sheets: dict[str, pd.DataFrame], sheet_name: str) -> None:
    """Mutate `handle` in-place to describe `sheet_name` instead of whatever
    sheet it currently describes. Used when the user switches sheet tabs in
    the preview -- no disk I/O required, the frame is already cached."""
    df = sheets[sheet_name]
    handle.active_sheet = sheet_name
    handle.row_count = len(df)
    handle.column_count = df.shape[1]
    handle.memory_usage_bytes = int(df.memory_usage(deep=True).sum())
    handle.column_profiles = profile_dataframe(df)
    handle.duplicate_row_count, handle.blank_cell_count = _row_level_stats(df)


def _load_delimited(
    path: Path, ext: str, large_file_threshold: int
) -> tuple[WorkbookHandle, dict[str, pd.DataFrame]]:
    sep = "\t" if ext == ".tsv" else ","

    with open(path, "rb") as fh:
        row_count_estimate = sum(1 for _ in fh) - 1  # minus header

    engine_used = "pandas"
    if row_count_estimate > large_file_threshold:
        logger.info(
            "{} has ~{} rows (> threshold {}), using Polars fast path.",
            path.name,
            row_count_estimate,
            large_file_threshold,
        )
        pl_df = pl.read_csv(path, separator=sep, infer_schema_length=10_000)
        df = pl_df.to_pandas()
        engine_used = "polars"
    else:
        df = pd.read_csv(path, sep=sep)

    handle = WorkbookHandle(
        file_path=path,
        display_name=path.name,
        extension=ext,
        engine_used=engine_used,  # type: ignore[arg-type]
        sheets=[SheetInfo(name="Sheet1", row_count=len(df), column_count=df.shape[1])],
        active_sheet="Sheet1",
        row_count=len(df),
        column_count=df.shape[1],
        memory_usage_bytes=int(df.memory_usage(deep=True).sum()),
        loaded_at=datetime.now(),
    )
    return handle, {"Sheet1": df}


def _load_with_pandas(
    path: Path, sheet_name: str | None, *, engine: str | None
) -> tuple[WorkbookHandle, dict[str, pd.DataFrame]]:
    visibility = {}
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            visibility = _sheet_names_and_visibility(path)
        except Exception:
            logger.warning("Could not read sheet visibility for {}", path.name, exc_info=True)

    all_sheets = pd.read_excel(path, sheet_name=None, engine=engine)
    active = sheet_name if sheet_name in all_sheets else next(iter(all_sheets))
    active_df = all_sheets[active]

    sheet_infos = [
        SheetInfo(
            name=name,
            row_count=len(sheet_df),
            column_count=sheet_df.shape[1],
            is_hidden=visibility.get(name, False),
            is_empty=sheet_df.empty,
        )
        for name, sheet_df in all_sheets.items()
    ]

    handle = WorkbookHandle(
        file_path=path,
        display_name=path.name,
        extension=path.suffix.lower(),
        engine_used="pandas",
        sheets=sheet_infos,
        active_sheet=active,
        row_count=len(active_df),
        column_count=active_df.shape[1],
        memory_usage_bytes=int(active_df.memory_usage(deep=True).sum()),
        loaded_at=datetime.now(),
    )
    return handle, all_sheets
